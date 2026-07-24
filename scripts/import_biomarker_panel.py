#!/usr/bin/env python3
"""Import selected_biomarker_target_list.xlsx into the committed panel JSON that
`preclinical_biomarker_surveyor` screens a case against.

Usage:
  python3 scripts/import_biomarker_panel.py [--workbook PATH] [--check]

Writes data/reference/selected_biomarker_panel.json. The workbook is the upstream
`cell_protein_target_paint` run's stage-3 output: one row per target on the
"Targets" sheet, one row per therapeutic binder on the "Binders" sheet, and the
provenance of the generating run on "Run Info".

This file is machine-generated and must never be hand-edited. The hand-curated
tumor-agnostic biomarkers (MSI/dMMR, TMB, and friends, which the workbook does not
cover because it enumerates protein targets rather than predictive biomarkers)
live separately in data/reference/tumor_agnostic_biomarkers.json so re-running
this importer cannot clobber them.

`--check` re-imports into memory and diffs against the committed file without
writing, so CI / a test can prove the committed JSON still matches the workbook.
Exit codes: 0 = written (or in sync, under --check), 1 = drift under --check,
2 = usage / missing input.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = REPO / "selected_biomarker_target_list.xlsx"
OUT_PATH = REPO / "data" / "reference" / "selected_biomarker_panel.json"

# Localization → what it actually takes to measure the target on a patient's
# tumor. Surface proteins are an IHC question; HLA-presented intracellular
# antigens additionally need a class I genotype before any TCR-T / ImmTAC is on
# the table, which is exactly the kind of gap this panel exists to surface.
ASSAY_BY_LOCALIZATION = {
    "surface": "Tumor-cell protein expression by IHC",
    "surface_and_secreted": "Tumor-cell protein expression by IHC; serum/plasma assay where a validated one exists",
    "intracellular_HLA_presented": "Antigen expression by IHC or RT-PCR, plus HLA class I genotyping",
}

# Placeholder used by the upstream workbook for a target with no binder program.
UNADDRESSED = "(unaddressed)"


def slugify(value: str) -> str:
    """Kebab-case key for a gene symbol, e.g. 'EGFR (variant III)' → 'egfr-variant-iii'."""
    out = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return out


def split_list(value, sep: str = ";") -> list[str]:
    if value is None:
        return []
    return [p.strip() for p in str(value).split(sep) if p and p.strip()]


def clean(value):
    """Normalize an openpyxl cell to str/int/None, collapsing blank strings to None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return text or None


def _rows(ws) -> list[dict]:
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return []
    header = [str(h).strip() if h is not None else "" for h in raw[0]]
    return [dict(zip(header, row)) for row in raw[1:]]


def _binder(row: dict) -> dict | None:
    name = clean(row.get("Binder Name"))
    if not name or name == UNADDRESSED:
        return None
    return {
        "binder_name": name,
        "modality": clean(row.get("Modality")),
        "sponsor": clean(row.get("Sponsor")),
        "development_stage": clean(row.get("Development Stage")),
        "highest_phase": clean(row.get("Highest Phase")),
        "indications": split_list(row.get("Indications")),
        "fda_approved": str(clean(row.get("FDA Approved?")) or "").lower() == "yes",
        "approval_year": clean(row.get("Approval Year")),
        "approved_indication": clean(row.get("Approved Indication")),
        "brand_name": clean(row.get("Brand Name")),
        "key_pmids": split_list(row.get("Key PMIDs")),
    }


def _as_int(value) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def build_panel(workbook: Path) -> dict:
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover - environment problem, not logic
        raise SystemExit(
            "import_biomarker_panel: missing dependency `openpyxl`.\n  pip install openpyxl"
        ) from e

    if not workbook.exists():
        raise SystemExit(f"import_biomarker_panel: no such workbook {workbook}")

    wb = openpyxl.load_workbook(workbook, data_only=True)
    for sheet in ("Targets", "Binders", "Run Info"):
        if sheet not in wb.sheetnames:
            raise SystemExit(f"import_biomarker_panel: workbook has no '{sheet}' sheet")

    run_info = {
        str(k).strip(): clean(v)
        for k, v in ws_pairs(wb["Run Info"])
        if k is not None and str(k).strip()
    }

    binders_by_gene: dict[str, list[dict]] = {}
    for row in _rows(wb["Binders"]):
        gene = clean(row.get("Target Gene Symbol"))
        binder = _binder(row)
        if gene and binder:
            binders_by_gene.setdefault(gene, []).append(binder)

    targets = []
    for row in _rows(wb["Targets"]):
        gene = clean(row.get("Gene Symbol"))
        if not gene:
            continue
        localization = clean(row.get("Localization"))
        targets.append(
            {
                "panel_key": slugify(gene),
                "gene_symbol": gene,
                "protein_name": clean(row.get("Protein Name")),
                "uniprot_id": clean(row.get("UniProt ID")),
                "aliases": split_list(row.get("Aliases")),
                "localization": localization,
                "localization_notes": clean(row.get("Localization Notes")),
                "best_evidence_level": clean(row.get("Best Evidence Level")),
                "cancer_relevance": clean(row.get("Cancer Relevance")),
                "tumor_vs_normal_evidence": clean(row.get("Tumor vs Normal Evidence")),
                "liabilities": clean(row.get("Liabilities")),
                "priority_hint": clean(row.get("Priority Hint")),
                "default_assay": ASSAY_BY_LOCALIZATION.get(
                    localization, "Tumor-cell protein expression by IHC"
                ),
                "evidence_pmids": split_list(row.get("Evidence PMIDs")),
                "binder_counts": {
                    "total": _as_int(row.get("# Binders")),
                    "fda_approved": _as_int(row.get("# FDA-Approved Binders")),
                    "clinical_stage": _as_int(row.get("# Clinical-Stage Binders")),
                    "preclinical": _as_int(row.get("# Preclinical Binders")),
                },
                "best_binder": (
                    None
                    if clean(row.get("Best Binder")) in (None, UNADDRESSED)
                    else {
                        "name": clean(row.get("Best Binder")),
                        "modality": clean(row.get("Best Binder Modality")),
                        "stage": clean(row.get("Best Binder Stage")),
                    }
                ),
                "binders": binders_by_gene.get(gene, []),
                "notes": clean(row.get("Notes")),
            }
        )

    targets.sort(key=lambda t: t["gene_symbol"])

    return {
        "panel_id": "selected-biomarker-target-list",
        "panel_title": "Selected biomarker target list",
        "description": (
            "Cell-surface and HLA-presented protein targets with an associated "
            "therapeutic binder program, screened against a Libby case by the "
            "preclinical_biomarker_surveyor agent. Machine-generated from the "
            "upstream cell_protein_target_paint workbook; do not hand-edit."
        ),
        "generator": "scripts/import_biomarker_panel.py",
        "source": {
            "workbook": workbook.name,
            "workbook_sha256": hashlib.sha256(workbook.read_bytes()).hexdigest(),
            "workflow": run_info.get("workflow"),
            "stage": run_info.get("stage"),
            "run_timestamp_iso": run_info.get("run_timestamp_iso"),
            "reference_audit": {
                "applied": str(run_info.get("reference_audit_applied") or "").lower() == "true",
                "refs_checked": _as_int(run_info.get("reference_audit.refs_checked")),
                "refs_passed": _as_int(run_info.get("reference_audit.refs_passed")),
                "refs_failed": _as_int(run_info.get("reference_audit.refs_failed")),
                "refs_flagged": _as_int(run_info.get("reference_audit.refs_flagged")),
                "strictness": run_info.get("reference_audit.strictness"),
            },
        },
        "target_count": len(targets),
        "targets": targets,
    }


def ws_pairs(ws):
    """Yield (key, value) from a two-column key/value sheet, skipping blank rows."""
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = row[0]
        value = row[1] if len(row) > 1 else None
        if key is None:
            continue
        yield key, value


def serialize(panel: dict) -> str:
    return json.dumps(panel, indent=2, ensure_ascii=False, sort_keys=False) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", type=Path, default=OUT_PATH)
    parser.add_argument(
        "--check",
        action="store_true",
        help="Verify the committed JSON matches the workbook; write nothing.",
    )
    args = parser.parse_args()

    panel = build_panel(args.workbook)
    text = serialize(panel)

    if args.check:
        if not args.out.exists():
            print(f"FAIL  {args.out} does not exist", file=sys.stderr)
            return 1
        committed = args.out.read_text(encoding="utf-8")
        # generated_utc is deliberately absent from the payload so this diff is
        # meaningful: the file changes only when the workbook's content changes.
        if committed != text:
            print(
                f"FAIL  {args.out.relative_to(REPO)} is out of sync with "
                f"{args.workbook.name}; re-run scripts/import_biomarker_panel.py",
                file=sys.stderr,
            )
            return 1
        print(f"OK    {args.out.relative_to(REPO)} matches {args.workbook.name}")
        return 0

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(text, encoding="utf-8")
    n_binders = sum(len(t["binders"]) for t in panel["targets"])
    print(
        f"wrote {args.out.relative_to(REPO)} "
        f"({panel['target_count']} targets, {n_binders} binders)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
